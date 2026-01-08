# -*- coding: utf-8 -*-
from __future__ import annotations
import sys
import os
import unicodedata
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from io import BytesIO
from dataclasses import dataclass
import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


@dataclass
class ContaContabil:
    codigo: str
    classificacao: str
    descricao: str
    saldo_atual: float
    tipo_conta: str
    eh_redutora: bool
    natureza_encontrada: str
    natureza_correta: str
    observacao: str = ''


def remover_acentos(texto):
    if not texto:
        return ''
    nfkd = unicodedata.normalize('NFKD', texto)
    return ''.join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()


def localizar_linha_cabecalho(ws):
    cabecalhos = ['codigo', 'classificacao', 'descricao da conta', 'saldo atual']
    for row_idx in range(1, min(ws.max_row + 1, 50)):
        valores = []
        for col_idx in range(1, ws.max_column + 1):
            valor = ws.cell(row=row_idx, column=col_idx).value
            if valor:
                valores.append(remover_acentos(str(valor)))
        if sum(1 for cab in cabecalhos if cab in valores) == len(cabecalhos):
            return row_idx
    return None


def identificar_colunas(ws, linha_cabecalho):
    colunas = {}
    mapeamento = {
        'codigo': 'codigo', 'classificacao': 'classificacao',
        'descricao da conta': 'descricao', 'saldo atual': 'saldo_atual'
    }
    for col_idx in range(1, ws.max_column + 1):
        valor = ws.cell(row=linha_cabecalho, column=col_idx).value
        if valor:
            valor_n = remover_acentos(str(valor))
            if valor_n in mapeamento:
                colunas[mapeamento[valor_n]] = col_idx
    return colunas


def verificar_celula_negrito(ws, row, col):
    celula = ws.cell(row=row, column=col)
    return celula.font and celula.font.bold


def eh_conta_analitica(ws, row, col_classificacao, classificacao):
    if verificar_celula_negrito(ws, row, col_classificacao):
        return False
    return True


def identificar_tipo_conta(classificacao):
    if not classificacao:
        return None
    mapeamento = {'1': 'Ativo', '2': 'Passivo', '3': 'Despesa', '4': 'Receita'}
    return mapeamento.get(classificacao.strip()[0])


def verificar_conta_redutora(descricao):
    return descricao.strip().startswith('(-)') if descricao else False


def determinar_natureza_encontrada(saldo_atual):
    if saldo_atual is None:
        return None
    if saldo_atual > 0:
        return 'D'
    elif saldo_atual < 0:
        return 'C'
    return None


def determinar_natureza_correta(tipo_conta, eh_redutora):
    if not tipo_conta:
        return None
    natureza_normal = {'Ativo': 'D', 'Passivo': 'C', 'Despesa': 'D', 'Receita': 'C'}
    natureza = natureza_normal.get(tipo_conta)
    if natureza and eh_redutora:
        natureza = 'C' if natureza == 'D' else 'D'
    return natureza


def gerar_observacao(tipo_conta, natureza_encontrada, eh_redutora):
    redutora_txt = 'Redutora de ' if eh_redutora else ''
    if natureza_encontrada == 'D':
        return redutora_txt + tipo_conta + ' com saldo devedor (esperado: credor)'
    else:
        return redutora_txt + tipo_conta + ' com saldo credor (esperado: devedor)'


def analisar_balancete_arquivo(arquivo_bytes):
    wb = load_workbook(BytesIO(arquivo_bytes), data_only=False)
    ws = wb['Balancete'] if 'Balancete' in wb.sheetnames else wb.active

    linha_cabecalho = localizar_linha_cabecalho(ws)
    if not linha_cabecalho:
        raise ValueError('Nao foi possivel localizar a linha de cabecalho.')

    colunas = identificar_colunas(ws, linha_cabecalho)
    obrigatorias = ['codigo', 'classificacao', 'descricao', 'saldo_atual']
    faltando = [col for col in obrigatorias if col not in colunas]
    if faltando:
        raise ValueError('Colunas nao encontradas: ' + str(faltando))

    contas_incorretas = []
    col_codigo = colunas['codigo']
    col_classificacao = colunas['classificacao']
    col_descricao = colunas['descricao']
    col_saldo_atual = colunas['saldo_atual']

    total_linhas = 0
    total_analiticas = 0
    total_saldo_zero = 0

    for row_idx in range(linha_cabecalho + 1, ws.max_row + 1):
        classificacao = ws.cell(row=row_idx, column=col_classificacao).value
        descricao = ws.cell(row=row_idx, column=col_descricao).value

        if not classificacao or not descricao:
            continue

        classificacao = str(classificacao).strip()
        descricao = str(descricao).strip()
        total_linhas += 1

        if not eh_conta_analitica(ws, row_idx, col_classificacao, classificacao):
            continue

        total_analiticas += 1
        codigo = ws.cell(row=row_idx, column=col_codigo).value
        codigo = str(codigo).strip() if codigo else ''
        saldo_atual = ws.cell(row=row_idx, column=col_saldo_atual).value

        try:
            if isinstance(saldo_atual, str):
                saldo_atual = saldo_atual.replace('.', '').replace(',', '.')
            saldo_atual = float(saldo_atual) if saldo_atual else 0.0
        except:
            saldo_atual = 0.0

        tipo_conta = identificar_tipo_conta(classificacao)
        if not tipo_conta:
            continue

        eh_redutora = verificar_conta_redutora(descricao)
        natureza_encontrada = determinar_natureza_encontrada(saldo_atual)

        if natureza_encontrada is None:
            total_saldo_zero += 1
            continue

        natureza_correta = determinar_natureza_correta(tipo_conta, eh_redutora)

        if natureza_encontrada != natureza_correta:
            contas_incorretas.append(ContaContabil(
                codigo=codigo, classificacao=classificacao, descricao=descricao,
                saldo_atual=saldo_atual, tipo_conta=tipo_conta, eh_redutora=eh_redutora,
                natureza_encontrada=natureza_encontrada, natureza_correta=natureza_correta,
                observacao=gerar_observacao(tipo_conta, natureza_encontrada, eh_redutora)
            ))

    return contas_incorretas, {
        'total_linhas': total_linhas, 'total_analiticas': total_analiticas,
        'total_saldo_zero': total_saldo_zero, 'total_incorretas': len(contas_incorretas)
    }


def gerar_relatorio_bytes(contas):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Natureza Incorreta'

    fonte_cab = Font(bold=True, color='FFFFFF')
    preenche_cab = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    borda = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    cabecalhos = ['Codigo', 'Classificacao', 'Descricao', 'Tipo', 'Saldo Atual', 'Nat.Encontrada', 'Nat.Correta', 'Redutora', 'Observacao']

    for col_idx, cab in enumerate(cabecalhos, 1):
        c = ws.cell(row=1, column=col_idx, value=cab)
        c.font = fonte_cab
        c.fill = preenche_cab
        c.border = borda
        c.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, conta in enumerate(contas, 2):
        codigo_fmt = str(conta.codigo).replace('.', '').replace(',', '').strip()
        dados = [codigo_fmt, conta.classificacao, conta.descricao, conta.tipo_conta, conta.saldo_atual,
                 conta.natureza_encontrada, conta.natureza_correta, 'Sim' if conta.eh_redutora else 'Nao', conta.observacao]
        for col_idx, valor in enumerate(dados, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=valor)
            c.border = borda
            if col_idx == 5:
                c.number_format = '#,##0.00'
            if col_idx == 1:
                c.number_format = '@'

    for col, w in {'A': 15, 'B': 18, 'C': 45, 'D': 12, 'E': 15, 'F': 14, 'G': 14, 'H': 10, 'I': 45}.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:I' + str(len(contas) + 1)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def fmt_valor(v):
    sinal = '-' if v < 0 else ''
    valor_abs = abs(v)
    inteiro = int(valor_abs)
    decimal = int(round((valor_abs - inteiro) * 100))
    inteiro_fmt = '{:,}'.format(inteiro).replace(',', '.')
    return 'R$ ' + sinal + inteiro_fmt + ',' + str(decimal).zfill(2)


def mostrar_pagina_auditoria_natureza():
    st.title('Auditoria de Natureza de Contas')
    st.caption('Analise automatica de contas contabeis com natureza incorreta no balancete')

    with st.expander('Ver regras de natureza contabil', expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            st.markdown('**Contas Normais:**')
            st.markdown('- Ativo (1.x) = Devedora (D)')
            st.markdown('- Passivo (2.x) = Credora (C)')
            st.markdown('- Despesa (3.x) = Devedora (D)')
            st.markdown('- Receita (4.x) = Credora (C)')
        with col2:
            st.markdown('**Contas Redutoras (-):**')
            st.markdown('- Redutora de Ativo = Credora (C)')
            st.markdown('- Redutora de Passivo = Devedora (D)')
            st.markdown('- Redutora de Despesa = Credora (C)')
            st.markdown('- Redutora de Receita = Devedora (D)')

    st.divider()

    arquivo_upload = st.file_uploader(
        'Selecione o arquivo de Balancete (.xlsx)',
        type=['xlsx'],
        help='O arquivo deve conter: Codigo, Classificacao, Descricao da conta, Saldo Atual'
    )

    if arquivo_upload is not None:
        try:
            with st.spinner('Analisando o balancete...'):
                arquivo_bytes = arquivo_upload.read()
                contas_incorretas, estatisticas = analisar_balancete_arquivo(arquivo_bytes)

            st.divider()

            c1, c2, c3, c4 = st.columns(4)
            c1.metric('Total de Contas', estatisticas['total_linhas'])
            c2.metric('Contas Analiticas', estatisticas['total_analiticas'])
            c3.metric('Saldo Zero', estatisticas['total_saldo_zero'])
            c4.metric('Com Erro', estatisticas['total_incorretas'])

            if contas_incorretas:
                st.divider()

                por_tipo = {}
                for conta in contas_incorretas:
                    if conta.tipo_conta not in por_tipo:
                        por_tipo[conta.tipo_conta] = {'qtd': 0, 'valor': 0}
                    por_tipo[conta.tipo_conta]['qtd'] += 1
                    por_tipo[conta.tipo_conta]['valor'] += abs(conta.saldo_atual)

                st.subheader('Distribuicao por Tipo')
                cols = st.columns(len(por_tipo))
                for i, (tipo, dados) in enumerate(sorted(por_tipo.items())):
                    with cols[i]:
                        st.metric(tipo, dados['qtd'], fmt_valor(dados['valor']))

                st.divider()

                st.subheader('Contas com Natureza Incorreta')
                col_filtro1, col_filtro2, col_filtro3 = st.columns([2, 2, 1])

                tipos_disponiveis = sorted(set(c.tipo_conta for c in contas_incorretas))

                with col_filtro1:
                    filtro_tipo = st.multiselect('Filtrar por tipo:', tipos_disponiveis, default=tipos_disponiveis)

                with col_filtro2:
                    filtro_redutora = st.selectbox('Contas redutoras:', ['Todas', 'Apenas redutoras', 'Apenas normais'])

                with col_filtro3:
                    ordenar_por = st.selectbox('Ordenar por:', ['Tipo', 'Saldo', 'Codigo'])

                contas_filtradas = [c for c in contas_incorretas if c.tipo_conta in filtro_tipo]

                if filtro_redutora == 'Apenas redutoras':
                    contas_filtradas = [c for c in contas_filtradas if c.eh_redutora]
                elif filtro_redutora == 'Apenas normais':
                    contas_filtradas = [c for c in contas_filtradas if not c.eh_redutora]

                if ordenar_por == 'Saldo':
                    contas_filtradas = sorted(contas_filtradas, key=lambda x: abs(x.saldo_atual), reverse=True)
                elif ordenar_por == 'Codigo':
                    contas_filtradas = sorted(contas_filtradas, key=lambda x: x.codigo)
                else:
                    contas_filtradas = sorted(contas_filtradas, key=lambda x: x.tipo_conta)

                st.write('**Exibindo ' + str(len(contas_filtradas)) + ' de ' + str(len(contas_incorretas)) + ' contas**')

                dados_tabela = []
                for conta in contas_filtradas:
                    codigo_fmt = str(conta.codigo).replace('.', '').replace(',', '').strip()
                    redutora_txt = ' (R)' if conta.eh_redutora else ''
                    dados_tabela.append({
                        'Tipo': conta.tipo_conta,
                        'Codigo': codigo_fmt,
                        'Classificacao': conta.classificacao,
                        'Descricao': conta.descricao + redutora_txt,
                        'Saldo': conta.saldo_atual,
                        'Nat. Atual': conta.natureza_encontrada,
                        'Nat. Correta': conta.natureza_correta
                    })

                df = pd.DataFrame(dados_tabela)

                st.dataframe(
                    df,
                    use_container_width=True,
                    hide_index=True,
                    height=min(400, 35 * len(df) + 38),
                    column_config={
                        'Tipo': st.column_config.TextColumn('Tipo', width='small'),
                        'Codigo': st.column_config.TextColumn('Codigo', width='small'),
                        'Classificacao': st.column_config.TextColumn('Classif.', width='medium'),
                        'Descricao': st.column_config.TextColumn('Descricao da Conta', width='large'),
                        'Saldo': st.column_config.NumberColumn('Saldo Atual', format='R$ %.2f'),
                        'Nat. Atual': st.column_config.TextColumn('Atual', width='small'),
                        'Nat. Correta': st.column_config.TextColumn('Correta', width='small')
                    }
                )

                st.caption('(R) = Conta Redutora')

                st.divider()

                st.subheader('Download do Relatorio')
                excel_bytes = gerar_relatorio_bytes(contas_incorretas)
                st.download_button(
                    label='Baixar Relatorio Excel',
                    data=excel_bytes,
                    file_name='Balancete_contas_natureza_incorreta.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    type='primary'
                )

            else:
                st.success('Nenhuma conta com natureza incorreta foi encontrada no balancete!')

        except ValueError as e:
            st.error('Erro na analise: ' + str(e))
        except Exception as e:
            st.error('Erro inesperado: ' + str(e))


if __name__ == '__main__':
    st.set_page_config(page_title='Auditoria de Natureza - VPS', layout='wide')
    mostrar_pagina_auditoria_natureza()
