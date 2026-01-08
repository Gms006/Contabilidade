# -*- coding: utf-8 -*-
"""
Módulo de Auditoria de Conciliação Bancária para VPS METALÚRGICA
Versão adaptada para Streamlit

Funcionalidades:
- Conciliação de lançamentos entre extrato e razão
- Extração do número do lançamento
- Extração do código reduzido da conta contábil
- Verificação cruzada entre bancos (lançamentos no banco errado)
"""

from __future__ import annotations

import re
import pandas as pd
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any
from io import BytesIO

# Tentar importar openpyxl para formatação Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# =============================================================================
# ESTILOS EXCEL
# =============================================================================

ESTILOS = {
    'cabecalho': {
        'font': Font(bold=True, color="FFFFFF", size=11),
        'fill': PatternFill(start_color="2D3E50", end_color="2D3E50", fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    },
    'ok': PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid"),
    'erro': PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid"),
    'sucesso_total': PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid"),
} if HAS_OPENPYXL else {}


# =============================================================================
# FUNÇÕES DE PARSING
# =============================================================================

def parse_valor(valor_raw: str) -> Tuple[float, str]:
    """Converte valor para float e identifica tipo (C/D)."""
    if pd.isna(valor_raw) or valor_raw is None:
        return 0.0, None
    
    valor_str = str(valor_raw).strip().upper()
    if not valor_str:
        return 0.0, None
    
    # Identificar tipo
    tipo = None
    if valor_str.endswith('C'):
        tipo, valor_str = 'C', valor_str[:-1]
    elif valor_str.endswith('D'):
        tipo, valor_str = 'D', valor_str[:-1]
    elif valor_str.startswith('-'):
        tipo, valor_str = 'D', valor_str[1:]
    
    # Limpar e converter
    valor_str = re.sub(r'[^\d,.\-]', '', valor_str)
    try:
        if '.' in valor_str and ',' in valor_str:
            valor_str = valor_str.replace('.', '').replace(',', '.')
        elif ',' in valor_str:
            valor_str = valor_str.replace(',', '.')
        valor_float = float(valor_str)
    except (ValueError, TypeError):
        return 0.0, None
    
    if tipo is None:
        tipo = 'C' if valor_float >= 0 else 'D'
        valor_float = abs(valor_float)
    
    return (-abs(valor_float) if tipo == 'D' else abs(valor_float)), tipo


def parse_data(data_raw) -> Optional[datetime]:
    """Converte data para datetime."""
    if pd.isna(data_raw) or data_raw is None:
        return None
    
    if isinstance(data_raw, (datetime, pd.Timestamp)):
        return pd.to_datetime(data_raw).normalize().to_pydatetime()
    
    data_str = str(data_raw).strip().split('\n')[0].strip()
    
    for fmt in ['%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d', '%d-%m-%Y']:
        try:
            return datetime.strptime(data_str, fmt)
        except ValueError:
            continue
    return None


def detectar_colunas(df: pd.DataFrame) -> Dict[str, Any]:
    """Detecta formato e colunas do arquivo."""
    # Formato simples: 3 colunas
    if df.shape[1] == 3:
        primeira = df.iloc[0, 0] if not df.empty else None
        if primeira and parse_data(primeira):
            return {'col_data': 0, 'col_hist': 1, 'col_valor': 2, 'linha_inicio': 0}
    
    # Procurar cabeçalho
    for idx, row in df.iterrows():
        linha = [str(v).lower().strip() if pd.notna(v) else '' for v in row]
        linha_str = ' '.join(linha)
        
        if 'data' in linha_str:
            cols = {}
            for i, val in enumerate(linha):
                if val == 'data': cols['col_data'] = i
                elif 'número' in val or 'numero' in val: 
                    # O número está 1 coluna à direita do cabeçalho devido a merge de células
                    cols['col_numero'] = i + 1
                elif 'histórico' in val or 'historico' in val: cols['col_hist'] = i
                elif 'valor' in val: cols['col_valor'] = i
                elif 'débito' in val or 'debito' in val: cols['col_debito'] = i
                elif 'crédito' in val or 'credito' in val: cols['col_credito'] = i
            
            if 'col_data' in cols:
                cols['linha_inicio'] = idx + 1
                return cols
    
    return {'col_data': 0, 'col_hist': 1, 'col_valor': 2, 'linha_inicio': 0}


# =============================================================================
# LEITURA DE ARQUIVOS
# =============================================================================

def ler_extrato_upload(arquivo, nome_banco: str = '') -> pd.DataFrame:
    """Lê arquivo de extrato bancário de upload do Streamlit."""
    try:
        xls = pd.ExcelFile(arquivo)
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet, header=None)
            if df.empty or len(df) < 2:
                continue
            
            cfg = detectar_colunas(df)
            registros = []
            
            for idx in range(cfg['linha_inicio'], len(df)):
                row = df.iloc[idx]
                data = parse_data(row.iloc[cfg['col_data']] if cfg['col_data'] < len(row) else None)
                if not data:
                    continue
                
                valor, tipo = parse_valor(row.iloc[cfg.get('col_valor', 2)] if cfg.get('col_valor', 2) < len(row) else None)
                if tipo is None:
                    continue
                
                hist = str(row.iloc[cfg.get('col_hist', 1)]).strip() if cfg.get('col_hist', 1) < len(row) and pd.notna(row.iloc[cfg.get('col_hist', 1)]) else ''
                hist = ' '.join(hist.split())
                
                if any(x in hist.lower() for x in ['saldo anterior', 'saldo do dia', 'saldo bloqueado']):
                    continue
                
                registros.append({
                    'data': data.date(), 
                    'historico': hist, 
                    'valor': valor, 
                    'tipo': tipo, 
                    'banco': nome_banco,
                    'origem': 'EXTRATO'
                })
            
            if registros:
                return pd.DataFrame(registros)
        return pd.DataFrame()
    except Exception as e:
        raise Exception(f"Erro ao ler extrato: {e}")


def extrair_info_conta_razao(df: pd.DataFrame, linha_cabecalho: int) -> Dict[str, str]:
    """
    Extrai informações da conta contábil do arquivo de razão.
    
    Estrutura padrão (linha após o cabeçalho):
    - Coluna 0: 'Conta:'
    - Coluna 2: Código reduzido da conta (7, 808, 809, etc.)
    - Coluna 9: Código completo (1.1.12.000.1, 1.1.12.000.4, etc.)
    - Coluna 15: Nome da conta (BANCO BRADESCO S.A., SICREDI, etc.)
    """
    info = {'codigo_contabil': '', 'codigo_reduzido': '', 'nome_conta': ''}
    
    for idx in range(linha_cabecalho, min(linha_cabecalho + 3, len(df))):
        row = df.iloc[idx]
        primeira_col = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        
        if 'conta:' in primeira_col or primeira_col == 'conta:':
            # Código reduzido está na coluna 2 (ex: 7, 808, 809)
            if len(row) > 2 and pd.notna(row.iloc[2]):
                codigo = row.iloc[2]
                if isinstance(codigo, float) and not pd.isna(codigo):
                    info['codigo_reduzido'] = str(int(codigo))
                else:
                    info['codigo_reduzido'] = str(codigo).strip()
            
            # Código completo está na coluna 9 (ex: 1.1.12.000.1)
            if len(row) > 9 and pd.notna(row.iloc[9]):
                codigo_completo = str(row.iloc[9]).strip()
                if re.match(r'\d+\.\d+\.\d+', codigo_completo):
                    info['codigo_contabil'] = codigo_completo
            
            # Nome da conta está na coluna 15 (ex: BANCO BRADESCO S.A.)
            if len(row) > 15 and pd.notna(row.iloc[15]):
                nome = str(row.iloc[15]).strip()
                if len(nome) > 2:
                    info['nome_conta'] = nome
            
            # Fallback: procurar em outras colunas se não encontrou
            if not info['nome_conta']:
                for col_idx in range(10, min(20, len(row))):
                    if pd.notna(row.iloc[col_idx]):
                        val = str(row.iloc[col_idx]).strip()
                        if len(val) > 3 and not val.replace('.', '').replace('-', '').isdigit():
                            info['nome_conta'] = val
                            break
            
            break
    
    return info


def ler_razao_upload(arquivo, nome_banco: str = '') -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Lê arquivo de razão contábil de upload do Streamlit."""
    try:
        xls = pd.ExcelFile(arquivo)
        sheet = xls.sheet_names[0]
        for nome in xls.sheet_names:
            if nome.lower() in ['razão', 'razao', 'balancete']:
                sheet = nome
                break
        
        df = pd.read_excel(xls, sheet_name=sheet, header=None)
        cfg = detectar_colunas(df)
        
        # Extrair info da conta com a função melhorada
        info = extrair_info_conta_razao(df, cfg.get('linha_inicio', 0))
        
        registros = []
        col_data = cfg.get('col_data', 0)
        col_numero = cfg.get('col_numero')  # Coluna do número do lançamento
        col_hist = cfg.get('col_hist')
        col_deb = cfg.get('col_debito')
        col_cred = cfg.get('col_credito')
        
        for idx in range(cfg.get('linha_inicio', 0), len(df)):
            row = df.iloc[idx]
            data = parse_data(row.iloc[col_data] if col_data < len(row) else None)
            if not data:
                continue
            
            hist = str(row.iloc[col_hist]).strip() if col_hist and col_hist < len(row) and pd.notna(row.iloc[col_hist]) else ''
            if any(x in hist.lower() for x in ['saldo anterior', 'saldo do dia', 'ajuste saldo']):
                continue
            
            # Extrair número do lançamento
            numero_lancamento = ''
            if col_numero and col_numero < len(row) and pd.notna(row.iloc[col_numero]):
                num_val = row.iloc[col_numero]
                if isinstance(num_val, float) and not pd.isna(num_val):
                    numero_lancamento = str(int(num_val))
                else:
                    numero_lancamento = str(num_val).strip()
            
            debito = float(row.iloc[col_deb]) if col_deb and col_deb < len(row) and pd.notna(row.iloc[col_deb]) else 0.0
            credito = float(row.iloc[col_cred]) if col_cred and col_cred < len(row) and pd.notna(row.iloc[col_cred]) else 0.0
            
            if debito == 0 and credito == 0:
                continue
            
            valor = debito - credito
            registros.append({
                'data': data.date(), 
                'numero': numero_lancamento,
                'historico': hist, 
                'valor': valor,
                'tipo': 'C' if valor >= 0 else 'D', 
                'banco': nome_banco,
                'origem': 'RAZAO'
            })
        
        return pd.DataFrame(registros), info
    except Exception as e:
        raise Exception(f"Erro ao ler razão: {e}")


# =============================================================================
# CONCILIAÇÃO
# =============================================================================

def conciliar_banco(df_extrato: pd.DataFrame, df_razao: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """
    Realiza conciliação entre extrato e razão de um banco.
    
    Retorna:
        - conciliacao: Todos os lançamentos com status
        - faltantes: No extrato mas NÃO contabilizados (precisam ser lançados)
        - indevidos: Contabilizados mas NÃO pertencem ao extrato (lançados errado)
    """
    if df_extrato.empty and df_razao.empty:
        return {'conciliacao': pd.DataFrame(), 'faltantes': pd.DataFrame(), 'indevidos': pd.DataFrame()}
    
    ext = df_extrato.copy() if not df_extrato.empty else pd.DataFrame()
    raz = df_razao.copy() if not df_razao.empty else pd.DataFrame()
    
    # Criar chave única para conciliação (data + valor arredondado + índice duplicata)
    for df in [ext, raz]:
        if not df.empty:
            df['valor_round'] = df['valor'].round(2)
            df['idx_dup'] = df.groupby(['data', 'valor_round']).cumcount()
            df['chave'] = df.apply(lambda x: f"{x['data']}_{x['valor_round']:.2f}_{x['idx_dup']}", axis=1)
    
    if ext.empty:
        raz['status'] = 'INDEVIDO'
        raz['erro_tipo'] = 'Lançado no sistema mas NÃO existe no extrato'
        return {'conciliacao': raz, 'faltantes': pd.DataFrame(), 'indevidos': raz}
    
    if raz.empty:
        ext['status'] = 'FALTANTE'
        ext['erro_tipo'] = 'No extrato mas NÃO contabilizado no sistema'
        return {'conciliacao': ext, 'faltantes': ext, 'indevidos': pd.DataFrame()}
    
    # Merge
    merged = ext.merge(raz, on='chave', how='outer', suffixes=('_ext', '_raz'), indicator=True)
    
    # Definir status
    merged['status'] = merged['_merge'].map({
        'both': 'OK',
        'left_only': 'FALTANTE',
        'right_only': 'INDEVIDO'
    })
    
    merged['erro_tipo'] = merged['status'].map({
        'OK': '',
        'FALTANTE': 'No extrato mas NÃO contabilizado no sistema',
        'INDEVIDO': 'Lançado no sistema mas NÃO existe no extrato'
    })
    
    # Consolidar colunas
    merged['data'] = merged['data_ext'].fillna(merged['data_raz'])
    merged['valor'] = merged['valor_round_ext'].fillna(merged['valor_round_raz'])
    merged['tipo'] = merged['tipo_ext'].fillna(merged['tipo_raz'])
    merged['historico_extrato'] = merged.get('historico_ext', '').fillna('')
    merged['historico_razao'] = merged.get('historico_raz', '').fillna('')
    
    # Número do lançamento (vem apenas do razão)
    if 'numero' in merged.columns:
        merged['numero_lancamento'] = merged['numero'].fillna('')
    elif 'numero_raz' in merged.columns:
        merged['numero_lancamento'] = merged['numero_raz'].fillna('')
    else:
        merged['numero_lancamento'] = ''
    
    # Preparar resultado
    cols = ['data', 'numero_lancamento', 'valor', 'tipo', 'historico_extrato', 'historico_razao', 'status', 'erro_tipo']
    conciliacao = merged[cols].copy()
    
    faltantes = conciliacao[conciliacao['status'] == 'FALTANTE'].copy()
    indevidos = conciliacao[conciliacao['status'] == 'INDEVIDO'].copy()
    
    return {'conciliacao': conciliacao, 'faltantes': faltantes, 'indevidos': indevidos}


# =============================================================================
# VERIFICAÇÃO CRUZADA ENTRE BANCOS
# =============================================================================

def verificar_cruzamento_bancos(dados_bancos: Dict[str, Dict]) -> pd.DataFrame:
    """
    Verifica se lançamentos foram contabilizados no banco errado.
    
    Para cada lançamento INDEVIDO de um banco, verifica se ele pertence
    ao extrato de outro banco (foi lançado na conta errada).
    
    Para cada lançamento FALTANTE de um banco, verifica se ele foi
    lançado no razão de outro banco.
    """
    cruzamentos = []
    chaves_processadas = set()  # Evitar duplicatas
    bancos = list(dados_bancos.keys())
    
    for banco in bancos:
        dados = dados_bancos[banco]
        indevidos = dados.get('indevidos', pd.DataFrame())
        faltantes = dados.get('faltantes', pd.DataFrame())
        
        # Verificar INDEVIDOS: lançamentos no razão deste banco que pertencem ao extrato de outro
        if not indevidos.empty:
            for _, row in indevidos.iterrows():
                valor = row['valor']
                data = row['data']
                
                for outro_banco in bancos:
                    if outro_banco == banco:
                        continue
                    
                    df_extrato_outro = dados_bancos[outro_banco].get('extrato', pd.DataFrame())
                    if df_extrato_outro.empty:
                        continue
                    
                    matches = df_extrato_outro[
                        (df_extrato_outro['data'] == data) & 
                        (abs(df_extrato_outro['valor'] - valor) < 0.01)
                    ]
                    
                    if not matches.empty:
                        chave = f"{data}_{valor:.2f}_{outro_banco}_{banco}"
                        if chave not in chaves_processadas:
                            chaves_processadas.add(chave)
                            cruzamentos.append({
                                'tipo_erro': 'LANÇADO NO BANCO ERRADO',
                                'banco_origem': outro_banco,
                                'banco_destino_errado': banco,
                                'data': data,
                                'valor': valor,
                                'numero_lancamento': row.get('numero_lancamento', ''),
                                'historico_extrato': matches.iloc[0].get('historico', ''),
                                'historico_razao': row.get('historico_razao', ''),
                                'explicacao': f"Pertence ao extrato {outro_banco} mas foi lançado no razão {banco}"
                            })
        
        # Verificar FALTANTES: lançamentos no extrato deste banco que foram lançados no razão de outro
        if not faltantes.empty:
            for _, row in faltantes.iterrows():
                valor = row['valor']
                data = row['data']
                
                for outro_banco in bancos:
                    if outro_banco == banco:
                        continue
                    
                    df_razao_outro = dados_bancos[outro_banco].get('razao', pd.DataFrame())
                    if df_razao_outro.empty:
                        continue
                    
                    matches = df_razao_outro[
                        (df_razao_outro['data'] == data) & 
                        (abs(df_razao_outro['valor'] - valor) < 0.01)
                    ]
                    
                    if not matches.empty:
                        chave = f"{data}_{valor:.2f}_{banco}_{outro_banco}"
                        if chave not in chaves_processadas:
                            chaves_processadas.add(chave)
                            cruzamentos.append({
                                'tipo_erro': 'LANÇADO NO BANCO ERRADO',
                                'banco_origem': banco,
                                'banco_destino_errado': outro_banco,
                                'data': data,
                                'valor': valor,
                                'numero_lancamento': matches.iloc[0].get('numero', ''),
                                'historico_extrato': row.get('historico_extrato', ''),
                                'historico_razao': matches.iloc[0].get('historico', ''),
                                'explicacao': f"Pertence ao extrato {banco} mas foi lançado no razão {outro_banco}"
                            })
    
    return pd.DataFrame(cruzamentos)


# =============================================================================
# GERAÇÃO DE RELATÓRIO EXCEL
# =============================================================================

def formatar_valor_br(valor: float, com_cifrao: bool = True) -> str:
    """Formata valor para padrão brasileiro (1.234,56)."""
    if pd.isna(valor) or valor is None:
        return ''
    valor_abs = abs(valor)
    valor_str = f"{valor_abs:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    if valor < 0:
        valor_str = f"-{valor_str}"
    if com_cifrao:
        return f"R$ {valor_str}"
    return valor_str


def formatar_data_br(data) -> str:
    """Formata data para DD/MM/AAAA."""
    if pd.isna(data) or data is None:
        return ''
    if isinstance(data, str):
        return data
    try:
        if hasattr(data, 'strftime'):
            return data.strftime('%d/%m/%Y')
        return str(data)
    except:
        return str(data)


def gerar_excel_auditoria(resultado_banco: Dict, info_conta: Dict, nome_banco: str) -> bytes:
    """Gera arquivo Excel com o relatório de auditoria de um banco."""
    
    buffer = BytesIO()
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Resumo
        resumo_data = {
            'Informação': [
                'Banco',
                'Código Reduzido',
                'Conta Contábil',
                'Nome da Conta',
                '',
                'Total Extrato',
                'Total Razão',
                'Conciliados (OK)',
                'Faltantes',
                'Indevidos'
            ],
            'Valor': [
                nome_banco,
                info_conta.get('codigo_reduzido', ''),
                info_conta.get('codigo_contabil', ''),
                info_conta.get('nome_conta', ''),
                '',
                len(resultado_banco.get('extrato', [])),
                len(resultado_banco.get('razao', [])),
                len(resultado_banco['conciliacao'][resultado_banco['conciliacao']['status'] == 'OK']) if not resultado_banco['conciliacao'].empty else 0,
                len(resultado_banco['faltantes']),
                len(resultado_banco['indevidos'])
            ]
        }
        df_resumo = pd.DataFrame(resumo_data)
        df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
        
        # Conciliação
        if not resultado_banco['conciliacao'].empty:
            df_conc = resultado_banco['conciliacao'].copy()
            df_conc['data'] = df_conc['data'].apply(formatar_data_br)
            df_conc.to_excel(writer, sheet_name='Conciliação', index=False)
        
        # Faltantes
        if not resultado_banco['faltantes'].empty:
            df_falt = resultado_banco['faltantes'].copy()
            df_falt['data'] = df_falt['data'].apply(formatar_data_br)
            df_falt.to_excel(writer, sheet_name='Faltantes', index=False)
        
        # Indevidos
        if not resultado_banco['indevidos'].empty:
            df_ind = resultado_banco['indevidos'].copy()
            df_ind['data'] = df_ind['data'].apply(formatar_data_br)
            df_ind.to_excel(writer, sheet_name='Indevidos', index=False)
    
    return buffer.getvalue()


def gerar_excel_cruzamento(df_cruzamentos: pd.DataFrame) -> bytes:
    """Gera arquivo Excel com o relatório de cruzamento entre bancos."""
    
    buffer = BytesIO()
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        if not df_cruzamentos.empty:
            df_cruz = df_cruzamentos.copy()
            df_cruz['data'] = df_cruz['data'].apply(formatar_data_br)
            df_cruz.to_excel(writer, sheet_name='Cruzamento Entre Bancos', index=False)
            
            # Resumo por banco
            agrupado = df_cruzamentos.groupby(['banco_origem', 'banco_destino_errado']).agg({
                'valor': ['count', 'sum']
            }).reset_index()
            agrupado.columns = ['Banco Correto', 'Banco Errado', 'Quantidade', 'Valor Total']
            agrupado.to_excel(writer, sheet_name='Resumo por Banco', index=False)
        else:
            pd.DataFrame({'Mensagem': ['Nenhum lançamento cruzado encontrado']}).to_excel(
                writer, sheet_name='Resultado', index=False
            )
    
    return buffer.getvalue()


# =============================================================================
# FUNÇÃO PRINCIPAL DE AUDITORIA
# =============================================================================

def executar_auditoria_completa(arquivos_bancos: Dict[str, Dict]) -> Dict:
    """
    Executa auditoria completa de todos os bancos.
    
    Args:
        arquivos_bancos: Dict com estrutura:
            {
                'BRADESCO': {'extrato': arquivo_extrato, 'razao': arquivo_razao},
                'SICOOB': {'extrato': arquivo_extrato, 'razao': arquivo_razao},
                'SICREDI': {'extrato': arquivo_extrato, 'razao': arquivo_razao},
            }
    
    Returns:
        Dict com resultados de cada banco e verificação cruzada
    """
    resultados = {}
    dados_bancos = {}
    
    for nome_banco, arquivos in arquivos_bancos.items():
        try:
            # Ler extrato
            df_extrato = ler_extrato_upload(arquivos['extrato'], nome_banco)
            
            # Ler razão
            df_razao, info_conta = ler_razao_upload(arquivos['razao'], nome_banco)
            
            # Filtrar razão pelo período do extrato
            if not df_razao.empty and not df_extrato.empty:
                data_min = df_extrato['data'].min()
                data_max = df_extrato['data'].max()
                df_razao = df_razao[(df_razao['data'] >= data_min) & (df_razao['data'] <= data_max)]
            
            # Conciliar
            resultado = conciliar_banco(df_extrato, df_razao)
            
            # Guardar para verificação cruzada
            dados_bancos[nome_banco] = {
                'extrato': df_extrato,
                'razao': df_razao,
                'faltantes': resultado['faltantes'],
                'indevidos': resultado['indevidos'],
                'info': info_conta
            }
            
            resultados[nome_banco] = {
                'extrato': df_extrato,
                'razao': df_razao,
                'conciliacao': resultado['conciliacao'],
                'faltantes': resultado['faltantes'],
                'indevidos': resultado['indevidos'],
                'info': info_conta,
                'stats': {
                    'total_extrato': len(df_extrato),
                    'total_razao': len(df_razao),
                    'ok': len(resultado['conciliacao'][resultado['conciliacao']['status'] == 'OK']) if not resultado['conciliacao'].empty else 0,
                    'faltantes': len(resultado['faltantes']),
                    'indevidos': len(resultado['indevidos'])
                }
            }
            
        except Exception as e:
            resultados[nome_banco] = {
                'erro': str(e),
                'stats': {'total_extrato': 0, 'total_razao': 0, 'ok': 0, 'faltantes': 0, 'indevidos': 0}
            }
    
    # Verificação cruzada entre bancos
    if len(dados_bancos) > 1:
        df_cruzamentos = verificar_cruzamento_bancos(dados_bancos)
        resultados['_cruzamento'] = df_cruzamentos
    else:
        resultados['_cruzamento'] = pd.DataFrame()
    
    return resultados
