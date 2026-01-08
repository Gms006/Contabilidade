# -*- coding: utf-8 -*-
"""
Página de Auditoria de Conciliação Bancária
Neto Contabilidade - Sistema de Conciliação

Audita lançamentos do sistema (Razão) vs Extrato Bancário
Identifica 2 tipos de erros:
    1. FALTANTES: Lançamentos do extrato NÃO contabilizados no sistema
    2. INDEVIDOS: Lançamentos contabilizados que NÃO pertencem ao extrato
"""

from __future__ import annotations

import re
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Optional, Tuple, Any

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# =============================================================================
# CORES DO TEMA
# =============================================================================
CORES = {
    "azul_escuro": "#2D3E50",
    "azul_medio": "#34495E",
    "dourado": "#C9A96E",
    "branco": "#FFFFFF",
    "cinza_claro": "#F5F7FA",
    "sucesso": "#27AE60",
    "erro": "#E74C3C",
    "aviso": "#F39C12",
}

ESTILOS_EXCEL = {
    'cabecalho': {
        'font': Font(bold=True, color="FFFFFF", size=11),
        'fill': PatternFill(start_color="2D3E50", end_color="2D3E50", fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    },
    'ok': PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid"),
    'erro': PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid"),
}


# =============================================================================
# FUNÇÕES DE PARSING
# =============================================================================

def parse_valor(valor_raw: str) -> Tuple[float, str]:
    """Converte valor para float e identifica tipo (C/D)."""
    if pd.isna(valor_raw) or valor_raw is None:
        return 0.0, None
    
    valor_str = str(valor_raw).strip().upper()
    if not valor_str:
        return 0.0, None
    
    tipo = None
    if valor_str.endswith('C'):
        tipo, valor_str = 'C', valor_str[:-1]
    elif valor_str.endswith('D'):
        tipo, valor_str = 'D', valor_str[:-1]
    elif valor_str.startswith('-'):
        tipo, valor_str = 'D', valor_str[1:]
    
    valor_str = re.sub(r'[^\d,.\-]', '', valor_str)
    try:
        if '.' in valor_str and ',' in valor_str:
            valor_str = valor_str.replace('.', '').replace(',', '.')
        elif ',' in valor_str:
            valor_str = valor_str.replace(',', '.')
        valor_float = float(valor_str)
    except (ValueError, TypeError):
        return 0.0, None
    
    if tipo is None:
        tipo = 'C' if valor_float >= 0 else 'D'
        valor_float = abs(valor_float)
    
    return (-abs(valor_float) if tipo == 'D' else abs(valor_float)), tipo


def parse_data(data_raw) -> Optional[datetime]:
    """Converte data para datetime."""
    if pd.isna(data_raw) or data_raw is None:
        return None
    
    if isinstance(data_raw, (datetime, pd.Timestamp)):
        return pd.to_datetime(data_raw).normalize().to_pydatetime()
    
    data_str = str(data_raw).strip().split('\n')[0].strip()
    
    for fmt in ['%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d', '%d-%m-%Y']:
        try:
            return datetime.strptime(data_str, fmt)
        except ValueError:
            continue
    return None


def detectar_colunas(df: pd.DataFrame) -> Dict[str, Any]:
    """Detecta formato e colunas do arquivo."""
    if df.shape[1] == 3:
        primeira = df.iloc[0, 0] if not df.empty else None
        if primeira and parse_data(primeira):
            return {'col_data': 0, 'col_hist': 1, 'col_valor': 2, 'linha_inicio': 0}
    
    for idx, row in df.iterrows():
        linha = [str(v).lower().strip() if pd.notna(v) else '' for v in row]
        linha_str = ' '.join(linha)
        
        if 'data' in linha_str:
            cols = {}
            for i, val in enumerate(linha):
                if val == 'data': cols['col_data'] = i
                elif 'histórico' in val or 'historico' in val: cols['col_hist'] = i
                elif 'valor' in val: cols['col_valor'] = i
                elif 'débito' in val or 'debito' in val: cols['col_debito'] = i
                elif 'crédito' in val or 'credito' in val: cols['col_credito'] = i
            
            if 'col_data' in cols:
                cols['linha_inicio'] = idx + 1
                return cols
    
    return {'col_data': 0, 'col_hist': 1, 'col_valor': 2, 'linha_inicio': 0}


# =============================================================================
# LEITURA DE ARQUIVOS
# =============================================================================

def ler_extrato(arquivo_bytes: bytes, nome_arquivo: str) -> pd.DataFrame:
    """Lê arquivo de extrato bancário."""
    try:
        xls = pd.ExcelFile(BytesIO(arquivo_bytes))
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet, header=None)
            if df.empty or len(df) < 2:
                continue
            
            cfg = detectar_colunas(df)
            registros = []
            
            for idx in range(cfg['linha_inicio'], len(df)):
                row = df.iloc[idx]
                data = parse_data(row.iloc[cfg['col_data']] if cfg['col_data'] < len(row) else None)
                if not data:
                    continue
                
                valor, tipo = parse_valor(row.iloc[cfg.get('col_valor', 2)] if cfg.get('col_valor', 2) < len(row) else None)
                if tipo is None:
                    continue
                
                hist = str(row.iloc[cfg.get('col_hist', 1)]).strip() if cfg.get('col_hist', 1) < len(row) and pd.notna(row.iloc[cfg.get('col_hist', 1)]) else ''
                hist = ' '.join(hist.split())
                
                if any(x in hist.lower() for x in ['saldo anterior', 'saldo do dia', 'saldo bloqueado']):
                    continue
                
                registros.append({
                    'data': data.date(),
                    'historico': hist,
                    'valor': valor,
                    'tipo': tipo,
                    'arquivo': nome_arquivo,
                    'origem': 'EXTRATO'
                })
            
            if registros:
                return pd.DataFrame(registros)
        return pd.DataFrame()
    except Exception as e:
        st.warning(f"Erro ao ler extrato {nome_arquivo}: {e}")
        return pd.DataFrame()


def ler_razao(arquivo_bytes: bytes, nome_arquivo: str) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Lê arquivo de razão contábil."""
    try:
        xls = pd.ExcelFile(BytesIO(arquivo_bytes))
        sheet = xls.sheet_names[0]
        for nome in xls.sheet_names:
            if nome.lower() in ['razão', 'razao', 'balancete']:
                sheet = nome
                break
        
        df = pd.read_excel(xls, sheet_name=sheet, header=None)
        cfg = detectar_colunas(df)
        
        # Extrair info da conta
        info = {'codigo_contabil': '', 'nome_conta': ''}
        for idx in range(cfg.get('linha_inicio', 0), min(cfg.get('linha_inicio', 0) + 5, len(df))):
            linha_str = ' '.join([str(v) if pd.notna(v) else '' for v in df.iloc[idx]])
            if 'conta:' in linha_str.lower():
                match = re.search(r'\d+\.\d+\.\d+\.\d+\.\d+', linha_str)
                if match:
                    info['codigo_contabil'] = match.group()
                for val in df.iloc[idx]:
                    val_str = str(val).strip() if pd.notna(val) else ''
                    if len(val_str) > 5 and not val_str.replace('.', '').replace('-', '').isdigit():
                        if 'conta:' not in val_str.lower() and len(val_str) > len(info['nome_conta']):
                            info['nome_conta'] = val_str
                break
        
        registros = []
        col_data = cfg.get('col_data', 0)
        col_hist = cfg.get('col_hist')
        col_deb = cfg.get('col_debito')
        col_cred = cfg.get('col_credito')
        
        for idx in range(cfg.get('linha_inicio', 0), len(df)):
            row = df.iloc[idx]
            data = parse_data(row.iloc[col_data] if col_data < len(row) else None)
            if not data:
                continue
            
            hist = str(row.iloc[col_hist]).strip() if col_hist and col_hist < len(row) and pd.notna(row.iloc[col_hist]) else ''
            if any(x in hist.lower() for x in ['saldo anterior', 'saldo do dia', 'ajuste saldo']):
                continue
            
            try:
                debito = float(row.iloc[col_deb]) if col_deb and col_deb < len(row) and pd.notna(row.iloc[col_deb]) else 0.0
            except:
                debito = 0.0
            try:
                credito = float(row.iloc[col_cred]) if col_cred and col_cred < len(row) and pd.notna(row.iloc[col_cred]) else 0.0
            except:
                credito = 0.0
            
            if debito == 0 and credito == 0:
                continue
            
            valor = debito - credito
            registros.append({
                'data': data.date(),
                'historico': hist,
                'valor': valor,
                'tipo': 'C' if valor >= 0 else 'D',
                'arquivo': nome_arquivo,
                'origem': 'RAZAO'
            })
        
        return pd.DataFrame(registros), info
    except Exception as e:
        st.warning(f"Erro ao ler razão {nome_arquivo}: {e}")
        return pd.DataFrame(), {}


# =============================================================================
# CONCILIAÇÃO
# =============================================================================

def conciliar(df_extrato: pd.DataFrame, df_razao: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """Realiza conciliação entre extrato e razão."""
    if df_extrato.empty and df_razao.empty:
        return {'conciliacao': pd.DataFrame(), 'faltantes': pd.DataFrame(), 'indevidos': pd.DataFrame()}
    
    ext = df_extrato.copy() if not df_extrato.empty else pd.DataFrame()
    raz = df_razao.copy() if not df_razao.empty else pd.DataFrame()
    
    for df in [ext, raz]:
        if not df.empty:
            df['valor_round'] = df['valor'].round(2)
            df['idx_dup'] = df.groupby(['data', 'valor_round']).cumcount()
            df['chave'] = df.apply(lambda x: f"{x['data']}_{x['valor_round']:.2f}_{x['idx_dup']}", axis=1)
    
    if ext.empty:
        raz['status'] = 'INDEVIDO'
        raz['erro_tipo'] = 'Lançado no sistema mas NÃO existe no extrato'
        return {'conciliacao': raz, 'faltantes': pd.DataFrame(), 'indevidos': raz}
    
    if raz.empty:
        ext['status'] = 'FALTANTE'
        ext['erro_tipo'] = 'No extrato mas NÃO contabilizado no sistema'
        return {'conciliacao': ext, 'faltantes': ext, 'indevidos': pd.DataFrame()}
    
    merged = ext.merge(raz, on='chave', how='outer', suffixes=('_ext', '_raz'), indicator=True)
    
    merged['status'] = merged['_merge'].map({
        'both': 'OK',
        'left_only': 'FALTANTE',
        'right_only': 'INDEVIDO'
    })
    
    merged['erro_tipo'] = merged['status'].map({
        'OK': '',
        'FALTANTE': 'No extrato mas NÃO contabilizado no sistema',
        'INDEVIDO': 'Lançado no sistema mas NÃO existe no extrato'
    })
    
    merged['data'] = merged['data_ext'].fillna(merged['data_raz'])
    merged['valor'] = merged['valor_round_ext'].fillna(merged['valor_round_raz'])
    merged['tipo'] = merged['tipo_ext'].fillna(merged['tipo_raz'])
    merged['historico_extrato'] = merged.get('historico_ext', '').fillna('')
    merged['historico_razao'] = merged.get('historico_raz', '').fillna('')
    merged['arquivo_extrato'] = merged.get('arquivo_ext', '').fillna('')
    merged['arquivo_razao'] = merged.get('arquivo_raz', '').fillna('')
    
    cols = ['data', 'valor', 'tipo', 'historico_extrato', 'historico_razao',
            'arquivo_extrato', 'arquivo_razao', 'status', 'erro_tipo']
    conciliacao = merged[cols].copy()
    
    faltantes = conciliacao[conciliacao['status'] == 'FALTANTE'].copy()
    indevidos = conciliacao[conciliacao['status'] == 'INDEVIDO'].copy()
    
    return {'conciliacao': conciliacao, 'faltantes': faltantes, 'indevidos': indevidos}


# =============================================================================
# FORMATAÇÃO
# =============================================================================

def formatar_valor_br(valor: float) -> str:
    """Formata valor para padrão brasileiro."""
    if pd.isna(valor) or valor is None:
        return ''
    valor_abs = abs(valor)
    valor_str = f"{valor_abs:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    if valor < 0:
        valor_str = f"-{valor_str}"
    return f"R$ {valor_str}"


def formatar_data_br(data) -> str:
    """Formata data para DD/MM/AAAA."""
    if pd.isna(data) or data is None:
        return ''
    if isinstance(data, str):
        return data
    try:
        if hasattr(data, 'strftime'):
            return data.strftime('%d/%m/%Y')
        return str(data)
    except:
        return str(data)


# =============================================================================
# GERAÇÃO DE RELATÓRIO EXCEL
# =============================================================================

def gerar_relatorio_excel(nome_banco: str, df_extrato: pd.DataFrame, df_razao: pd.DataFrame,
                          resultado: Dict, info_conta: Dict) -> bytes:
    """Gera arquivo XLSX com 3 abas: Resumo, Conciliação, Erros."""
    
    wb = Workbook()
    
    data_min = df_extrato['data'].min() if not df_extrato.empty else None
    data_max = df_extrato['data'].max() if not df_extrato.empty else None
    qtd_ok = len(resultado['conciliacao'][resultado['conciliacao']['status'] == 'OK']) if not resultado['conciliacao'].empty else 0
    qtd_faltantes = len(resultado['faltantes'])
    qtd_indevidos = len(resultado['indevidos'])
    total_erros = qtd_faltantes + qtd_indevidos
    is_100_correto = total_erros == 0
    
    # ==================== ABA RESUMO ====================
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    ws_resumo['A1'] = f"AUDITORIA CONCILIAÇÃO BANCÁRIA - {nome_banco}"
    ws_resumo['A1'].font = Font(bold=True, size=16, color="2D3E50")
    ws_resumo.merge_cells('A1:D1')
    
    ws_resumo['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_resumo['A2'].font = Font(italic=True, color="666666")
    
    ws_resumo['A4'] = "STATUS DA AUDITORIA"
    ws_resumo['A4'].font = Font(bold=True, size=12, color="C9A96E")
    
    if is_100_correto:
        ws_resumo['A5'] = "✅ 100% CORRETO - Nenhum erro encontrado!"
        ws_resumo['A5'].font = Font(bold=True, size=14, color="27AE60")
        ws_resumo['A5'].fill = ESTILOS_EXCEL['ok']
    else:
        ws_resumo['A5'] = f"⚠️ ERROS ENCONTRADOS: {total_erros} lançamentos"
        ws_resumo['A5'].font = Font(bold=True, size=14, color="E74C3C")
        ws_resumo['A5'].fill = ESTILOS_EXCEL['erro']
    ws_resumo.merge_cells('A5:D5')
    
    info_rows = [
        ('Conta Contábil:', info_conta.get('codigo_contabil', '')),
        ('Nome da Conta:', info_conta.get('nome_conta', '')),
        ('Período Auditado:', f"{formatar_data_br(data_min)} a {formatar_data_br(data_max)}" if data_min else "N/A"),
    ]
    for i, (label, valor) in enumerate(info_rows, start=7):
        ws_resumo[f'A{i}'] = label
        ws_resumo[f'A{i}'].font = Font(bold=True)
        ws_resumo[f'B{i}'] = valor
    
    ws_resumo['A11'] = "ESTATÍSTICAS"
    ws_resumo['A11'].font = Font(bold=True, size=12, color="C9A96E")
    
    valor_ext = df_extrato['valor'].sum() if not df_extrato.empty else 0
    valor_raz = df_razao['valor'].sum() if not df_razao.empty else 0
    
    stats = [
        ('Lançamentos no Extrato:', len(df_extrato)),
        ('Lançamentos no Razão:', len(df_razao)),
        ('Conciliados (OK):', qtd_ok),
        ('', ''),
        ('ERROS - Faltantes (não contabilizados):', qtd_faltantes),
        ('ERROS - Indevidos (não pertencem):', qtd_indevidos),
        ('', ''),
        ('Valor Total Extrato:', formatar_valor_br(valor_ext)),
        ('Valor Total Razão:', formatar_valor_br(valor_raz)),
        ('Diferença:', formatar_valor_br(valor_ext - valor_raz)),
    ]
    
    for i, (label, valor) in enumerate(stats, start=12):
        ws_resumo[f'A{i}'] = label
        ws_resumo[f'A{i}'].font = Font(bold=True)
        ws_resumo[f'B{i}'] = valor
        if 'ERROS' in str(label) and isinstance(valor, int) and valor > 0:
            ws_resumo[f'B{i}'].font = Font(color="E74C3C", bold=True)
            ws_resumo[f'B{i}'].fill = ESTILOS_EXCEL['erro']
    
    ws_resumo.column_dimensions['A'].width = 45
    ws_resumo.column_dimensions['B'].width = 40
    
    # ==================== ABA CONCILIAÇÃO ====================
    ws_conc = wb.create_sheet("Conciliação")
    
    if not resultado['conciliacao'].empty:
        headers = ['Data', 'Valor', 'Tipo', 'Histórico Extrato', 'Histórico Razão', 'Status']
        for col, header in enumerate(headers, start=1):
            cell = ws_conc.cell(row=1, column=col, value=header)
            for key, val in ESTILOS_EXCEL['cabecalho'].items():
                setattr(cell, key, val)
        
        df_conc = resultado['conciliacao']
        for row_idx, row in enumerate(df_conc.itertuples(), start=2):
            ws_conc.cell(row=row_idx, column=1, value=formatar_data_br(row.data))
            cell_val = ws_conc.cell(row=row_idx, column=2, value=round(float(row.valor), 2) if pd.notna(row.valor) else 0)
            cell_val.number_format = '#,##0.00'
            ws_conc.cell(row=row_idx, column=3, value=row.tipo)
            ws_conc.cell(row=row_idx, column=4, value=row.historico_extrato if pd.notna(row.historico_extrato) else '')
            ws_conc.cell(row=row_idx, column=5, value=row.historico_razao if pd.notna(row.historico_razao) else '')
            ws_conc.cell(row=row_idx, column=6, value=row.status)
            
            fill = ESTILOS_EXCEL['ok'] if row.status == 'OK' else ESTILOS_EXCEL['erro']
            for col in range(1, 7):
                ws_conc.cell(row=row_idx, column=col).fill = fill
        
        for i, w in enumerate([12, 15, 8, 50, 50, 15], start=1):
            ws_conc.column_dimensions[chr(64+i)].width = w
        ws_conc.auto_filter.ref = f"A1:F{len(df_conc)+1}"
    
    # ==================== ABA ERROS ====================
    ws_erros = wb.create_sheet("Erros")
    
    faltantes = resultado['faltantes']
    indevidos = resultado['indevidos']
    
    if faltantes.empty and indevidos.empty:
        ws_erros['A1'] = "✅ AUDITORIA 100% CORRETA"
        ws_erros['A1'].font = Font(bold=True, size=16, color="27AE60")
        ws_erros.merge_cells('A1:E1')
        ws_erros['A3'] = "Nenhum erro encontrado na conciliação bancária."
        ws_erros['A5'] = "• Todos os lançamentos do extrato foram contabilizados corretamente."
        ws_erros['A6'] = "• Não existem lançamentos indevidos no sistema."
        for i in [1, 3, 5, 6]:
            ws_erros[f'A{i}'].fill = ESTILOS_EXCEL['ok']
    else:
        row_atual = 1
        
        ws_erros[f'A{row_atual}'] = "❌ LANÇAMENTOS FALTANTES (não contabilizados)"
        ws_erros[f'A{row_atual}'].font = Font(bold=True, size=12, color="E74C3C")
        ws_erros.merge_cells(f'A{row_atual}:E{row_atual}')
        row_atual += 1
        
        ws_erros[f'A{row_atual}'] = "Estes lançamentos estão no EXTRATO mas NÃO foram lançados no sistema."
        ws_erros[f'A{row_atual}'].font = Font(italic=True, size=10, color="666666")
        row_atual += 1
        
        if not faltantes.empty:
            headers = ['Data', 'Valor', 'Tipo', 'Histórico', 'Arquivo']
            for col, header in enumerate(headers, start=1):
                cell = ws_erros.cell(row=row_atual, column=col, value=header)
                for key, val in ESTILOS_EXCEL['cabecalho'].items():
                    setattr(cell, key, val)
            row_atual += 1
            
            for _, row in faltantes.iterrows():
                ws_erros.cell(row=row_atual, column=1, value=formatar_data_br(row['data']))
                cell_val = ws_erros.cell(row=row_atual, column=2, value=round(float(row['valor']), 2) if pd.notna(row['valor']) else 0)
                cell_val.number_format = '#,##0.00'
                ws_erros.cell(row=row_atual, column=3, value=row['tipo'])
                ws_erros.cell(row=row_atual, column=4, value=row['historico_extrato'] if pd.notna(row['historico_extrato']) else '')
                ws_erros.cell(row=row_atual, column=5, value=row['arquivo_extrato'] if pd.notna(row['arquivo_extrato']) else '')
                for col in range(1, 6):
                    ws_erros.cell(row=row_atual, column=col).fill = ESTILOS_EXCEL['erro']
                row_atual += 1
        else:
            ws_erros[f'A{row_atual}'] = "Nenhum lançamento faltante ✓"
            ws_erros[f'A{row_atual}'].font = Font(color="27AE60")
            row_atual += 1
        
        row_atual += 2
        
        ws_erros[f'A{row_atual}'] = "❌ LANÇAMENTOS INDEVIDOS (não pertencem ao extrato)"
        ws_erros[f'A{row_atual}'].font = Font(bold=True, size=12, color="E74C3C")
        ws_erros.merge_cells(f'A{row_atual}:E{row_atual}')
        row_atual += 1
        
        ws_erros[f'A{row_atual}'] = "Estes lançamentos foram contabilizados mas NÃO existem no extrato."
        ws_erros[f'A{row_atual}'].font = Font(italic=True, size=10, color="666666")
        row_atual += 1
        
        if not indevidos.empty:
            headers = ['Data', 'Valor', 'Tipo', 'Histórico', 'Arquivo']
            for col, header in enumerate(headers, start=1):
                cell = ws_erros.cell(row=row_atual, column=col, value=header)
                for key, val in ESTILOS_EXCEL['cabecalho'].items():
                    setattr(cell, key, val)
            row_atual += 1
            
            for _, row in indevidos.iterrows():
                ws_erros.cell(row=row_atual, column=1, value=formatar_data_br(row['data']))
                cell_val = ws_erros.cell(row=row_atual, column=2, value=round(float(row['valor']), 2) if pd.notna(row['valor']) else 0)
                cell_val.number_format = '#,##0.00'
                ws_erros.cell(row=row_atual, column=3, value=row['tipo'])
                ws_erros.cell(row=row_atual, column=4, value=row['historico_razao'] if pd.notna(row['historico_razao']) else '')
                ws_erros.cell(row=row_atual, column=5, value=row['arquivo_razao'] if pd.notna(row['arquivo_razao']) else '')
                for col in range(1, 6):
                    ws_erros.cell(row=row_atual, column=col).fill = ESTILOS_EXCEL['erro']
                row_atual += 1
        else:
            ws_erros[f'A{row_atual}'] = "Nenhum lançamento indevido ✓"
            ws_erros[f'A{row_atual}'].font = Font(color="27AE60")
    
    for i, w in enumerate([12, 15, 8, 60, 30], start=1):
        ws_erros.column_dimensions[chr(64+i)].width = w
    
    # Salvar em bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# =============================================================================
# INTERFACE STREAMLIT
# =============================================================================

def mostrar_pagina_auditoria_bancaria():
    """Página principal de auditoria de conciliação bancária."""
    
    st.markdown("""
    <h1 style="color: #2D3E50; border-bottom: 3px solid #C9A96E; padding-bottom: 10px;">
        🏦 Auditoria de Conciliação Bancária
    </h1>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <div style="background: linear-gradient(135deg, #f5f7fa 0%, #e8ecf0 100%); 
                padding: 20px; border-radius: 10px; margin-bottom: 20px;
                border-left: 4px solid #C9A96E;">
        <p style="margin: 0; color: #2D3E50;">
            <strong>📋 Objetivo:</strong> Auditar lançamentos do sistema (Razão) vs Extrato Bancário<br>
            <strong>🔍 Identifica:</strong> Lançamentos faltantes e lançamentos indevidos
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Inicializar session state
    if 'bancos_auditoria' not in st.session_state:
        st.session_state.bancos_auditoria = []
    
    # ==========================================================================
    # SEÇÃO: CONFIGURAR BANCOS
    # ==========================================================================
    
    st.markdown("### 📁 Configurar Bancos para Auditoria")
    
    with st.expander("➕ Adicionar Novo Banco", expanded=len(st.session_state.bancos_auditoria) == 0):
        col1, col2 = st.columns([1, 2])
        
        with col1:
            nome_banco = st.text_input(
                "Nome do Banco",
                placeholder="Ex: BRADESCO, SICOOB, ITAÚ...",
                help="Digite o nome do banco para identificação"
            )
        
        with col2:
            st.markdown("##### 📤 Upload dos Arquivos")
        
        col_ext, col_raz = st.columns(2)
        
        with col_ext:
            st.markdown("""
            <div style="background: #e8f4f8; padding: 10px; border-radius: 8px; margin-bottom: 10px;">
                <strong>📄 Extratos Bancários</strong><br>
                <small style="color: #666;">Pode enviar múltiplos arquivos (um por mês)</small>
            </div>
            """, unsafe_allow_html=True)
            arquivos_extrato = st.file_uploader(
                "Extratos",
                type=['xlsx', 'xls'],
                accept_multiple_files=True,
                key="upload_extrato",
                label_visibility="collapsed"
            )
        
        with col_raz:
            st.markdown("""
            <div style="background: #f8f4e8; padding: 10px; border-radius: 8px; margin-bottom: 10px;">
                <strong>📊 Razão Contábil</strong><br>
                <small style="color: #666;">Arquivo do razão da conta bancária</small>
            </div>
            """, unsafe_allow_html=True)
            arquivo_razao = st.file_uploader(
                "Razão",
                type=['xlsx', 'xls'],
                accept_multiple_files=False,
                key="upload_razao",
                label_visibility="collapsed"
            )
        
        if st.button("➕ Adicionar Banco", type="primary", disabled=not nome_banco):
            if nome_banco and (arquivos_extrato or arquivo_razao):
                banco_config = {
                    'nome': nome_banco.upper(),
                    'extratos': [(f.name, f.read()) for f in arquivos_extrato] if arquivos_extrato else [],
                    'razao': (arquivo_razao.name, arquivo_razao.read()) if arquivo_razao else None
                }
                st.session_state.bancos_auditoria.append(banco_config)
                st.success(f"✅ Banco {nome_banco.upper()} adicionado com sucesso!")
                st.rerun()
            else:
                st.warning("⚠️ Adicione pelo menos um arquivo de extrato ou razão.")
    
    # ==========================================================================
    # EXIBIR BANCOS CONFIGURADOS
    # ==========================================================================
    
    if st.session_state.bancos_auditoria:
        st.markdown("### 🏦 Bancos Configurados")
        
        for idx, banco in enumerate(st.session_state.bancos_auditoria):
            with st.container():
                col1, col2, col3, col4 = st.columns([2, 3, 3, 1])
                
                with col1:
                    st.markdown(f"""
                    <div style="background: #2D3E50; color: white; padding: 10px 15px; 
                                border-radius: 8px; text-align: center; font-weight: bold;">
                        🏦 {banco['nome']}
                    </div>
                    """, unsafe_allow_html=True)
                
                with col2:
                    qtd_ext = len(banco['extratos'])
                    st.markdown(f"📄 **Extratos:** {qtd_ext} arquivo(s)")
                
                with col3:
                    tem_razao = "✅ Sim" if banco['razao'] else "❌ Não"
                    st.markdown(f"📊 **Razão:** {tem_razao}")
                
                with col4:
                    if st.button("🗑️", key=f"remove_{idx}", help="Remover banco"):
                        st.session_state.bancos_auditoria.pop(idx)
                        st.rerun()
        
        st.markdown("---")
        
        # ==========================================================================
        # BOTÃO EXECUTAR AUDITORIA
        # ==========================================================================
        
        col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
        with col_btn2:
            executar = st.button(
                "🔍 EXECUTAR AUDITORIA",
                type="primary",
                use_container_width=True
            )
        
        if executar:
            st.markdown("---")
            st.markdown("## 📊 Resultados da Auditoria")
            
            resultados_gerais = []
            arquivos_download = {}
            
            for banco in st.session_state.bancos_auditoria:
                with st.spinner(f"Processando {banco['nome']}..."):
                    
                    # Carregar extratos
                    dfs_extrato = []
                    for nome_arq, bytes_arq in banco['extratos']:
                        df = ler_extrato(bytes_arq, nome_arq)
                        if not df.empty:
                            dfs_extrato.append(df)
                    
                    df_extrato = pd.concat(dfs_extrato, ignore_index=True).sort_values('data').reset_index(drop=True) if dfs_extrato else pd.DataFrame()
                    
                    # Carregar razão
                    if banco['razao']:
                        nome_raz, bytes_raz = banco['razao']
                        df_razao, info_conta = ler_razao(bytes_raz, nome_raz)
                    else:
                        df_razao = pd.DataFrame()
                        info_conta = {}
                    
                    # Filtrar razão pelo período do extrato
                    if not df_extrato.empty and not df_razao.empty:
                        data_min, data_max = df_extrato['data'].min(), df_extrato['data'].max()
                        df_razao = df_razao[(df_razao['data'] >= data_min) & (df_razao['data'] <= data_max)]
                    
                    # Conciliar
                    resultado = conciliar(df_extrato, df_razao)
                    
                    # Estatísticas
                    qtd_ok = len(resultado['conciliacao'][resultado['conciliacao']['status'] == 'OK']) if not resultado['conciliacao'].empty else 0
                    qtd_faltantes = len(resultado['faltantes'])
                    qtd_indevidos = len(resultado['indevidos'])
                    total_erros = qtd_faltantes + qtd_indevidos
                    is_correto = total_erros == 0
                    
                    resultados_gerais.append({
                        'banco': banco['nome'],
                        'extrato': len(df_extrato),
                        'razao': len(df_razao),
                        'ok': qtd_ok,
                        'faltantes': qtd_faltantes,
                        'indevidos': qtd_indevidos,
                        'is_correto': is_correto
                    })
                    
                    # Gerar Excel
                    excel_bytes = gerar_relatorio_excel(banco['nome'], df_extrato, df_razao, resultado, info_conta)
                    arquivos_download[banco['nome']] = excel_bytes
                    
                    # ==========================================================================
                    # EXIBIR RESULTADO DO BANCO
                    # ==========================================================================
                    
                    st.markdown(f"### 🏦 {banco['nome']}")
                    
                    # Métricas
                    col1, col2, col3, col4, col5 = st.columns(5)
                    
                    with col1:
                        st.metric("📄 Extrato", len(df_extrato))
                    with col2:
                        st.metric("📊 Razão", len(df_razao))
                    with col3:
                        st.metric("✅ Conciliados", qtd_ok)
                    with col4:
                        st.metric("⚠️ Faltantes", qtd_faltantes, delta=None if qtd_faltantes == 0 else f"-{qtd_faltantes}", delta_color="inverse")
                    with col5:
                        st.metric("❌ Indevidos", qtd_indevidos, delta=None if qtd_indevidos == 0 else f"-{qtd_indevidos}", delta_color="inverse")
                    
                    # Status
                    if is_correto:
                        st.success("✅ **100% CORRETO** - Nenhum erro encontrado!")
                    else:
                        st.error(f"⚠️ **ERROS ENCONTRADOS:** {total_erros} lançamentos precisam de atenção")
                        
                        # Tabs para ver detalhes
                        tab1, tab2 = st.tabs(["⚠️ Faltantes", "❌ Indevidos"])
                        
                        with tab1:
                            if not resultado['faltantes'].empty:
                                st.markdown("**Lançamentos no extrato NÃO contabilizados:**")
                                df_show = resultado['faltantes'][['data', 'valor', 'tipo', 'historico_extrato']].copy()
                                df_show.columns = ['Data', 'Valor', 'Tipo', 'Histórico']
                                df_show['Data'] = df_show['Data'].apply(formatar_data_br)
                                df_show['Valor'] = df_show['Valor'].apply(lambda x: f"{x:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
                                st.dataframe(df_show, use_container_width=True, hide_index=True)
                            else:
                                st.info("Nenhum lançamento faltante ✓")
                        
                        with tab2:
                            if not resultado['indevidos'].empty:
                                st.markdown("**Lançamentos contabilizados que NÃO existem no extrato:**")
                                df_show = resultado['indevidos'][['data', 'valor', 'tipo', 'historico_razao']].copy()
                                df_show.columns = ['Data', 'Valor', 'Tipo', 'Histórico']
                                df_show['Data'] = df_show['Data'].apply(formatar_data_br)
                                df_show['Valor'] = df_show['Valor'].apply(lambda x: f"{x:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
                                st.dataframe(df_show, use_container_width=True, hide_index=True)
                            else:
                                st.info("Nenhum lançamento indevido ✓")
                    
                    # Download
                    st.download_button(
                        label=f"📥 Baixar Relatório {banco['nome']}.xlsx",
                        data=excel_bytes,
                        file_name=f"auditoria_{banco['nome']}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    st.markdown("---")
            
            # ==========================================================================
            # RESUMO GERAL
            # ==========================================================================
            
            st.markdown("## 📈 Resumo Geral da Auditoria")
            
            total_faltantes = sum(r['faltantes'] for r in resultados_gerais)
            total_indevidos = sum(r['indevidos'] for r in resultados_gerais)
            todos_corretos = all(r['is_correto'] for r in resultados_gerais)
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("🏦 Bancos Auditados", len(resultados_gerais))
            with col2:
                st.metric("⚠️ Total Faltantes", total_faltantes)
            with col3:
                st.metric("❌ Total Indevidos", total_indevidos)
            
            if todos_corretos:
                st.balloons()
                st.success("🎉 **PARABÉNS!** Todas as conciliações estão 100% corretas!")
            else:
                st.warning(f"⚠️ **ATENÇÃO:** {total_faltantes + total_indevidos} lançamentos precisam ser verificados.")
            
            # Tabela resumo
            df_resumo = pd.DataFrame(resultados_gerais)
            df_resumo.columns = ['Banco', 'Extrato', 'Razão', 'OK', 'Faltantes', 'Indevidos', 'Status']
            df_resumo['Status'] = df_resumo['Status'].apply(lambda x: '✅ OK' if x else '⚠️ Erros')
            st.dataframe(df_resumo, use_container_width=True, hide_index=True)
    
    else:
        st.info("👆 Adicione pelo menos um banco para iniciar a auditoria.")


# Para importação no app.py
if __name__ == "__main__":
    mostrar_pagina_auditoria_bancaria()
